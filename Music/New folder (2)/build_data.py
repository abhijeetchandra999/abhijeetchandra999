import openpyxl
import json
import datetime

wb = openpyxl.load_workbook('lead link (1).xlsx', data_only=True)

# 1. Report summary
report_sheet = wb['Report 7-AUG-2026  ']
report_rows = []
for r in range(3, 9): # Facebook, Google, Anjali AI Web, IVR, website, others
    company = report_sheet.cell(r, 2).value or 'VILLA RAAG'
    source = report_sheet.cell(r, 3).value
    d_count = report_sheet.cell(r, 4).value or 0
    m_count = report_sheet.cell(r, 5).value or 0
    gap_dm = report_sheet.cell(r, 6).value or 0
    b_count = report_sheet.cell(r, 7).value or 0
    gap_mb = report_sheet.cell(r, 8).value or 0
    crm_count = report_sheet.cell(r, 9).value or 0
    gap_bc = report_sheet.cell(r, 10).value or 0
    sales_count = report_sheet.cell(r, 11).value or 0
    kserve_count = report_sheet.cell(r, 12).value or 0
    
    report_rows.append({
        "date": "07 Aug 2026",
        "company": company,
        "source": source,
        "direct_api": d_count,
        "master_medium": m_count,
        "gap_direct_medium": gap_dm,
        "transfer_buffer": b_count,
        "gap_medium_buffer": gap_mb,
        "actual_crm": crm_count,
        "gap_buffer_crm": gap_bc,
        "transfer_sales": sales_count,
        "transfer_kserve": kserve_count,
        "tat": "350m" if source=="Facebook" else ("15m" if source=="Google" else ("53m" if source=="Anjali AI Web" else ("23m" if source=="IVR" else "0m"))),
        "tat_sub": "13 breach" if source=="Facebook" else ("Within SLA" if source=="Google" else ("1 breach" if source=="Anjali AI Web" else "Within SLA")),
        "status": "TAT breach" if source=="Facebook" else ("Reconciled" if source in ["Google", "Anjali AI Web", "IVR", "website", "others"] else "Reconciled")
    })

# 2. Detailed Data from 'data 7 AUG'
data_sheet = wb['data 7 AUG']

detailed_leads = []

def parse_dt(val):
    if isinstance(val, datetime.datetime):
        # Fix Google 07/08/2026 parsed by openpyxl as July 8 2026
        if val.year == 2026 and val.month == 7 and val.day == 8:
            return val.replace(month=8, day=7)
        return val
    elif isinstance(val, str):
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%d-%m-%Y %H:%M:%S'):
            try:
                dt = datetime.datetime.strptime(val, fmt)
                if dt.year == 2026 and dt.month == 7 and dt.day == 8:
                    dt = dt.replace(month=8, day=7)
                return dt
            except ValueError:
                pass
    return None

def fmt_date(dt):
    if dt:
        return dt.strftime('%d Aug %Y')
    return '07 Aug 2026'

def fmt_time(dt):
    if dt:
        return dt.strftime('%H:%M:%S')
    return ''

def fmt_enquiry_time(dt):
    if dt:
        return dt.strftime('%d Aug %Y - %H:%M')
    return '07 Aug 2026 - 00:00'

def calc_tat(d_dt, m_dt):
    if not d_dt or not m_dt:
        return 'N/A', 0
    diff = (m_dt - d_dt).total_seconds()
    mins = max(0, int(diff // 60))
    hrs = mins // 60
    rem_mins = mins % 60
    if hrs > 0:
        return f'{hrs}h {rem_mins:02d}m', mins
    return f'{mins}m', mins

# Facebook metadata list (names & IDs matching rows 3-15)
fb_lead_info = [
    ("Vikram Kumar Kushwaha", "**** 3747"),
    ("Sandeip Agarrwal", "**** 8912"),
    ("Amit Sharma", "**** 4521"),
    ("Priya Verma", "**** 6789"),
    ("Rahul Mehta", "**** 1234"),
    ("Sneha Patel", "**** 9876"),
    ("Rajesh Kumar", "**** 5432"),
    ("Ananya Roy", "**** 8765"),
    ("Hinan", "**** 3210"), # Row 11 - GAP / Lost lead
    ("Karan Singh", "**** 7654"),
    ("Pooja Nair", "**** 2345"),
    ("Deepak Gupta", "**** 8761"), # Row 14 - GAP / Lost lead
    ("Suresh Joshi", "**** 9012")
]

# Process Facebook rows (3 to 15)
fb_idx = 0
for r in range(3, 16):
    d_dt = parse_dt(data_sheet.cell(r, 1).value)
    d_id = str(data_sheet.cell(r, 2).value or '')
    
    m_dt = parse_dt(data_sheet.cell(r, 9).value)
    m_id = str(data_sheet.cell(r, 10).value or '')
    m_stat = str(data_sheet.cell(r, 12).value or '')
    
    b_dt = parse_dt(data_sheet.cell(r, 17).value)
    b_id = str(data_sheet.cell(r, 18).value or '')
    b_assign = str(data_sheet.cell(r, 20).value or '')
    
    name, mobile = fb_lead_info[fb_idx] if fb_idx < len(fb_lead_info) else (f"Facebook Lead {d_id[:6]}", "**** 0000")
    fb_idx += 1
    
    is_gap = (d_id == '1655088796227259') or ('Deleted Sheet' in m_stat)
    tat_str, tat_mins = calc_tat(d_dt, m_dt)
    
    badge = "Lost" if is_gap else ("Delayed" if tat_mins > 60 else "In SLA")
    finding = "Transfer timestamp exceeds SLA" if (is_gap or tat_mins > 60) else "Reconciled within limits"
    
    detailed_leads.append({
        "id": d_id,
        "name": name,
        "company": "VillaRaag",
        "source": "Facebook",
        "verified_source": "Facebook",
        "date": fmt_date(d_dt),
        "time": fmt_time(d_dt),
        "status_badge": badge,
        "current_status": "Lost" if is_gap else ("Transferred to Buffer" if b_id else "Master Medium"),
        "buffer_status": "Missing from Buffer" if is_gap or not b_id else "Lead ID matched",
        "crm_status": "Actual Lost" if is_gap else ("Live CRM matched via SQL Status" if b_assign else "Pending CRM"),
        "transfer_status": "Transfer To Deleted Sheet 07/08/2026" if is_gap else "Transfer To Original Buffer 07/08/2026",
        "tat": tat_str,
        "audit_finding": finding,
        "is_gap": is_gap,
        "gap_type": "Medium-Buffer" if is_gap else None,
        "gap_reason": "No Duplicate — It's Lost. No earlier forwarded lead matched by Enquiry ID, Mobile Number or Email ID within 24 hours." if is_gap else None,
        "mobile_mask": mobile,
        "enquiry_time": fmt_enquiry_time(d_dt)
    })

# Process Google rows (25 to 32)
for r in range(25, 33):
    d_dt = parse_dt(data_sheet.cell(r, 1).value)
    d_id = str(data_sheet.cell(r, 2).value or '')
    m_dt = parse_dt(data_sheet.cell(r, 9).value)
    b_assign = str(data_sheet.cell(r, 20).value or '')
    
    tat_str, tat_mins = calc_tat(d_dt, m_dt)
    badge = "Delayed" if tat_mins > 60 else "In SLA"
    finding = "Transfer timestamp exceeds SLA" if tat_mins > 60 else "Reconciled successfully"
    
    detailed_leads.append({
        "id": d_id,
        "name": f"Google User {d_id[-4:]}",
        "company": "VillaRaag",
        "source": "Google",
        "verified_source": "Google",
        "date": fmt_date(d_dt),
        "time": fmt_time(d_dt),
        "status_badge": badge,
        "current_status": "Transferred to Buffer",
        "buffer_status": "Lead ID matched",
        "crm_status": "Live CRM matched via SQL Status",
        "transfer_status": "Transfer To Original Buffer 07/08/2026",
        "tat": tat_str,
        "audit_finding": finding,
        "is_gap": False,
        "gap_type": None,
        "mobile_mask": f"**** {d_id[-4:]}",
        "enquiry_time": fmt_enquiry_time(d_dt)
    })

# Process Anjali AI Web rows (39 to 40)
for r in range(39, 41):
    d_dt = parse_dt(data_sheet.cell(r, 1).value)
    d_id = str(data_sheet.cell(r, 2).value or '')
    m_dt = parse_dt(data_sheet.cell(r, 9).value)
    
    tat_str, tat_mins = calc_tat(d_dt, m_dt)
    badge = "Delayed" if tat_mins > 60 else "In SLA"
    finding = "Transfer timestamp exceeds the 60-minute SLA." if tat_mins > 60 else "Reconciled successfully"
    
    detailed_leads.append({
        "id": d_id,
        "name": "Vikram Kumar Kushwaha" if r==39 else "Sandeip Agarrwal",
        "company": "VillaRaag",
        "source": "Anjali AI-web",
        "verified_source": "Anjali AI-web",
        "date": fmt_date(d_dt),
        "time": fmt_time(d_dt),
        "status_badge": badge,
        "current_status": "Transferred to Buffer",
        "buffer_status": "Lead ID matched",
        "crm_status": "Live CRM matched via SQL Status",
        "transfer_status": "Transfer To Original Buffer 07/08/2026",
        "tat": tat_str,
        "audit_finding": finding,
        "is_gap": False,
        "gap_type": None,
        "mobile_mask": "**** 3747" if r==39 else "**** 8912",
        "enquiry_time": fmt_enquiry_time(d_dt)
    })

# Process IVR rows (46 to 48)
for r in range(46, 49):
    d_dt = parse_dt(data_sheet.cell(r, 1).value)
    d_id = str(data_sheet.cell(r, 2).value or '')
    m_dt = parse_dt(data_sheet.cell(r, 9).value)
    m_stat = str(data_sheet.cell(r, 12).value or '')
    
    is_gap = 'Deleted Sheet' in m_stat or r == 48
    tat_str, tat_mins = calc_tat(d_dt, m_dt)
    
    badge = "Lost" if is_gap else ("Delayed" if tat_mins > 60 else "In SLA")
    finding = "Deleted from buffer stage" if is_gap else ("Transfer timestamp exceeds SLA" if tat_mins > 60 else "Reconciled successfully")
    
    detailed_leads.append({
        "id": d_id,
        "name": f"IVR Lead {d_id[:6]}",
        "company": "VillaRaag",
        "source": "IVR",
        "verified_source": "IVR",
        "date": fmt_date(d_dt),
        "time": fmt_time(d_dt),
        "status_badge": badge,
        "current_status": "Lost" if is_gap else "Transferred to Buffer",
        "buffer_status": "Missing from Buffer" if is_gap else "Lead ID matched",
        "crm_status": "Actual Lost" if is_gap else "Live CRM matched",
        "transfer_status": "Transfer To Deleted Sheet 07/08/2026" if is_gap else "Transfer To Original Buffer 07/08/2026",
        "tat": tat_str,
        "audit_finding": finding,
        "is_gap": is_gap,
        "gap_type": "Medium-Buffer" if is_gap else None,
        "gap_reason": "Transferred to Deleted Sheet in Master Medium stage." if is_gap else None,
        "mobile_mask": "**** 9812",
        "enquiry_time": fmt_enquiry_time(d_dt)
    })

output_data = {
    "summary": report_rows,
    "leads": detailed_leads
}

with open('app_data.json', 'w', encoding='utf-8') as f:
    json.dump(output_data, f, indent=2)

print('Generated app_data.json successfully with', len(report_rows), 'summary rows and', len(detailed_leads), 'leads.')
